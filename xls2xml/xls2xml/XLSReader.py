"""
This module reads an Excel file and allows the user to get all the valid worksheet names,
get the 1st line in a worksheet and iterate over the rest of the worksheet row by row
(next_row). The returned row is a hash which contains only the keys that are defined in
a configuration file.

This module depends on openpyxl and pyyaml.
"""

from openpyxl import load_workbook
import yaml

WORKSHEETS_KEY_NAME = 'worksheets'
REQUIRED_HEADERS_KEY_NAME = 'required'
OPTIONAL_HEADERS_KEY_NAME = 'optional'

class XLSReader(object):
    """
    Reader for Excel file for the fields from worksheets defined in a configuration file

    Methods
    -------
    valid_worksheets
        return the list of valid worksheet names in the Excel file

    get_headers_by_worksheet
        return the list of field names in the first row of a worksheet.

    next_row
        return a data row in a worksheet as a hash for all the fields defined in the
        configuration file.

    """

    def __init__(self, xls_filename, conf_filename):
        """
        Constructor

        :param xls_filename: Excel file path
        :type xls_filename: basestring
        :param conf_filename: configuration file path
        :type conf_filename: basestring
        """
        with open(conf_filename, 'r') as conf_file:
            self.xls_conf = yaml.load(conf_file)
        self.workbook = load_workbook(xls_filename, read_only=True)
        self.worksheets = None
        self.row_offset = {}
        self.headers = {}

    def valid_worksheets(self):
        """
        :return: list of valid worksheet names in the Excel file
        :rtype: list
        """
        if self.worksheets is not None:
            return self.worksheets

        self.worksheets = []
        sheet_titles = self.workbook.sheetnames
        for title in self.xls_conf[WORKSHEETS_KEY_NAME]:
            # Check worksheet exists
            if title not in sheet_titles:
                continue

            # Check number of rows
            worksheet = self.workbook[title]
            if worksheet.max_row < 2:
                continue

            # Check required headers are present
            if title not in self.headers:
                self.headers[title] = []
            for cell in worksheet[1]:
                header_value = cell.value
                if header_value is None:
                    self.headers[title].append(header_value)
                else:
                    self.headers[title].append(header_value.strip())
            required_headers = self.xls_conf[title][REQUIRED_HEADERS_KEY_NAME]
            required_header_not_found = False
            for required_header in required_headers:
                if required_header not in self.headers[title]:
                    required_header_not_found = True
                    break
            if required_header_not_found:
                continue

            self.worksheets.append(title)

        return self.worksheets

    def get_headers_by_worksheet(self, worksheet):
        """
        :param worksheet: the name of the worksheet
        :type worksheet: basestring
        :return: the list of valid worksheet names in the Excel file
        :rtype: list
        """
        worksheets = self.valid_worksheets()
        if worksheet not in worksheets:
            print 'Worksheet '+worksheet+' is not available or not valid!'
            return []

        return [x for x in self.headers[worksheet] if x is not None]


    def next_row(self, worksheet):
        """
        Retrieve next data row

        :param worksheet: the name of the worksheet
        :type worksheet: basestring
        :return: A hash containing all the REQUIRED and OPTIONAL fields as keys
                and the corresponding data as values
        :rtype: dict
        """
        if self.worksheets is None:
            self.valid_worksheets()

        if worksheet not in self.worksheets:
            print 'Worksheet ' + worksheet + ' is not valid!'
            return False

        if worksheet not in self.row_offset:
            self.row_offset[worksheet] = 1
        self.row_offset[worksheet] += 1

        work_sheet = self.workbook[worksheet]
        required_headers = self.xls_conf[worksheet][REQUIRED_HEADERS_KEY_NAME]
        optional_headers = self.xls_conf[worksheet][OPTIONAL_HEADERS_KEY_NAME]

        for row in work_sheet.iter_rows(min_row=self.row_offset[worksheet]):
            num_cells = 0
            for cell in row:
                num_cells += 1

            data = {}
            go_to_next_row = False
            has_notnull = False
            for header in required_headers+optional_headers:
                header_index = self.headers[worksheet].index(header)
                if header_index >= num_cells:
                    go_to_next_row = True
                    break

                cell = row[header_index]
                if cell.value is not None:
                    has_notnull = True

                if isinstance(cell.value, unicode):
                    data[header] = cell.value.encode('ascii', 'ignore')
                else:
                    data[header] = cell.value

            if go_to_next_row or not has_notnull:
                self.row_offset[worksheet] += 1
                continue

            data['row_num'] = self.row_offset[worksheet]
            return data

        return False
